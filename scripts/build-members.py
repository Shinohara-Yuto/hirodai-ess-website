#!/usr/bin/env python3
"""メンバー紹介.xlsx から members.json を生成します。"""
from __future__ import annotations

import json
import re
import shutil
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("openpyxl が必要です: pip install openpyxl")
    sys.exit(1)

ROOT = Path(__file__).resolve().parents[2]
XLSX = ROOT / "メンバー紹介.xlsx"
PHOTO_SRC = ROOT / "メンバー紹介写真"
WEBSITE = ROOT / "website"
PHOTO_DST = WEBSITE / "assets" / "members"
JSON_OUT = WEBSITE / "data" / "members.json"

SECTION_GRADES = {
    "三年生": 3,
    "3年生": 3,
    "二年生": 2,
    "2年生": 2,
    "一年生": 1,
    "1年生": 1,
}


def parse_photo(value) -> int | None:
    if value is None or str(value).strip() == "":
        return None
    text = str(value).strip()
    match = re.search(r"(\d+)", text)
    if match:
        return int(match.group(1))
    return None


def cell_text(row: list, index: int) -> str:
    if index >= len(row) or row[index] is None:
        return ""
    return str(row[index]).strip()


def is_section_row(first_cell: str) -> int | None:
    return SECTION_GRADES.get(first_cell.strip())


def is_header_row(row: list) -> bool:
    first = cell_text(row, 0)
    if first in SECTION_GRADES:
        return cell_text(row, 1) in ("学部", "名前", "name")
    return False


def copy_photos() -> None:
    PHOTO_DST.mkdir(parents=True, exist_ok=True)
    if not PHOTO_SRC.exists():
        print(f"警告: 写真フォルダが見つかりません: {PHOTO_SRC}")
        return
    for path in PHOTO_SRC.iterdir():
        if path.suffix.lower() in {".jpg", ".jpeg", ".png", ".webp"}:
            shutil.copy2(path, PHOTO_DST / path.name)


def read_members_from_excel() -> list[dict]:
    if not XLSX.exists():
        raise FileNotFoundError(f"Excel が見つかりません: {XLSX}")

    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb.active

    members: list[dict] = []
    current_grade: int | None = None

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column, values_only=True):
        if not any(row):
            continue

        first = cell_text(row, 0)
        section_grade = is_section_row(first)
        if section_grade is not None:
            current_grade = section_grade
            if is_header_row(row):
                continue
            continue

        if current_grade is None or not first:
            continue

        members.append(
            {
                "grade": current_grade,
                "name": first,
                "faculty": cell_text(row, 1),
                "hometown": cell_text(row, 2),
                "hobby": cell_text(row, 3),
                "photo": parse_photo(row[4] if len(row) > 4 else None),
            }
        )

    return members


def main() -> None:
    copy_photos()
    members = read_members_from_excel()

    if not members:
        print("警告: Excel にデータがありません。")
        sys.exit(1)

    members.sort(key=lambda m: (
        -(m["grade"] or 0),
        m.get("photo") if m.get("photo") is not None else 999,
        m.get("name", ""),
    ))

    counts = {3: 0, 2: 0, 1: 0}
    for member in members:
        if member["grade"] in counts:
            counts[member["grade"]] += 1

    payload = {
        "source": XLSX.name,
        "counts": {
            "grade3": counts[3],
            "grade2": counts[2],
            "grade1": counts[1],
            "total": len(members),
        },
        "members": members,
    }

    JSON_OUT.parent.mkdir(parents=True, exist_ok=True)
    JSON_OUT.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"生成完了: {JSON_OUT} ({len(members)} 件)")
    print(f"  3年生: {counts[3]} / 2年生: {counts[2]} / 1年生: {counts[1]}")


if __name__ == "__main__":
    main()
